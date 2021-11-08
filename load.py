from openpyxl import load_workbook
wb = load_workbook('加班数据.xlsx')
# print(wb.sheetnames)
for sheet in  wb:
  print(sheet.title)
  print(sheet.max_row)
  print(sheet.max_column)
  print(sheet['F2'].value)

template_xlsx = 'working hours 模板.xlsx'
wb = load_workbook(template_xlsx)
# print(wb.sheetnames)
for sheet in  wb:
  print(sheet.title)
  print(sheet.max_row)
  print(sheet.max_column)
  print(sheet['F2'].value)
  sheet['F2'] = 'Test write'

wb.save(template_xlsx)
print(wb['Sheet1']['F2'].value)
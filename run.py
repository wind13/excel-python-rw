from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from helper import find_in_range, log

user='jennifer'
day='2021-10'

log('start...')
# ot_xlsx = '加班数据-9月.xlsx'
# ot_xlsx = './2021/DBB OT request1105.xlsx'
ot_xlsx = './2021/' + user + '-ot-' + day + '.xlsx'
ot_default_sheet_name = 'Sheet1'
# ot_sheet_name = 'OT'
#TODO 考虑通过查找相应的列名来定位具体的列： Resource ID, Start Date, End Date, Actual Hours involved
ot_staff_id_column = 'C'
ot_hours_column = 'J'
ot_start_date_column = 'K'
ot_end_date_column = 'L'
ot_start_row = 2

ot_wb = load_workbook(ot_xlsx)
log('ot loaded...')
n_ot_sheets = len(ot_wb.worksheets)
# log('n_ot_sheets...' + str(n_ot_sheets))
# log('first sheet name: ' + ot_wb.worksheets[0].title)
ot_sheet0_name = ot_wb.worksheets[0].title
ot_sheet_name = ot_sheet0_name if n_ot_sheets == 1 else ot_default_sheet_name
# log('ot_sheet_name:' + ot_sheet_name)
ot_ws = ot_wb[ot_sheet_name]
ot_end_row = ot_ws.max_row + 1

# log(ot_end_row) # 178

# wh_xlsx = '9月billing人力-01.xlsx'
# wh_xlsx = './2021/DBB-Oct supplier billing record for HSBC Audit.xlsx'
# wh_xlsx = './2021/Jennifer-2021-10-Billing-HR.xlsx'
wh_xlsx = './2021/' + user + '-staff-' + day + '.xlsx'
wh_default_sheet_name = 'Sheet1'
wh_staff_id_column = 'B'
wh_start_row = 3
# wh_holiday_color = 'FFC4BD97'

wh_wb = load_workbook(wh_xlsx)
n_wh_sheets = len(wh_wb.worksheets)
wh_sheet0_name = wh_wb.worksheets[0].title
wh_sheet_name = wh_sheet0_name if n_wh_sheets == 1 else wh_default_sheet_name
# log(wh_sheet_name)
wh_ws = wh_wb[wh_sheet_name]

# log(str(wh_ws.max_row)) # 378
# log(str(wh_ws.max_column)) # 45

wh_end_row = wh_ws.max_row + 1
wh_end_column = wh_ws.max_column + 1
wh_end_column_letter = get_column_letter(wh_ws.max_column)

for row in range(ot_start_row, ot_end_row):
  staff_id = ot_ws[ot_staff_id_column + str(row)].value
  # log(staff_id)
  hours = ot_ws[ot_hours_column + str(row)].value
  # log(hours)
  start_day = ot_ws[ot_start_date_column + str(row)].value
  # log(start_day)
  end_day = ot_ws[ot_end_date_column + str(row)].value
  # log(end_day)

  range_sid_from = wh_staff_id_column + str(wh_start_row) # C3
  range_sid_to = wh_staff_id_column + str(wh_end_row) # C379
  wh_staff_range = wh_ws[range_sid_from +  ':' + range_sid_to]
  staff_cell = find_in_range(staff_id, wh_staff_range)

  if staff_cell == None:
    # log('这个 ' + str(staff_id) + ' 在人力表中没找到对应的行。')
    continue

  range_day_from = 'A1'
  range_day_to = wh_end_column_letter + '1'
  wh_day_range = wh_ws[range_day_from +  ':' + range_day_to]
  day_cell = find_in_range(start_day, wh_day_range)

  if day_cell == None:
    # log('这个 ' + str(start_day) + ' 在人力表中没找到对应的日期。')
    continue

  # log(staff_cell.row)
  # log(day_cell.column_letter)
  sid_day_cell = wh_ws[day_cell.column_letter + str(staff_cell.row)]
  sid_day_value = sid_day_cell.value
  # log(day_cell.fill.fgColor)
  # log(sid_day_value)
  # log(hours)
  # is_holiday = day_cell.fill.fgColor.rgb != None and day_cell.fill.fgColor.rgb == wh_holiday_color
  # h = hours if is_holiday else (8 + hours)
  h = sid_day_value + hours
  wh_ws[day_cell.column_letter + str(staff_cell.row)] = h
  comment = Comment('extended service ' + str(hours) + 'h', ' ')
  wh_ws[day_cell.column_letter + str(staff_cell.row)].comment = comment

  # log(sid_day_cell.value)

# wh_wb.save('./2021/ruby-2021-1108-ot_hours.xlsx')
wh_wb.save('./2021/' + user + '-result-' + day + '.xlsx')

# log(ot_ws[ot_end_date_column + str(ot_end_row)].column)
# log(ot_ws[ot_end_date_column + str(ot_end_row)].column_letter)


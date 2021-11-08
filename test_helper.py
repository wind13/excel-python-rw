#!/usr/bin/python
# -*- coding: utf-8 -*-

import unittest
import datetime
from helper import find_in_range
from openpyxl import load_workbook


class TestHelper(unittest.TestCase):

  test_xlsx = 'test_data.xlsx'
  test_wb = load_workbook(test_xlsx)
  test_ws = test_wb['Sheet1']

  @classmethod
  def setUpClass(cls):
    print("this setupclass() method only called once.\n")

  def test_find_sheets(self):
    """Test find sheets"""
    self.assertEqual(1, len(self.test_wb.worksheets))

  def test_find_in_range(self):
    """Test find_in_range(range)"""
    range_staff_ids = self.test_ws['F2:F178']
    cell_44084211 = find_in_range(44084211, range_staff_ids)
    self.assertEqual(4, cell_44084211.row)
    cell_unkown = find_in_range(798734, range_staff_ids)
    self.assertEqual(None, cell_unkown)

  def test_find_date(self):
    """Test find_date(range)"""
    range_start_dates = self.test_ws['O2:O178']
    the_day = datetime.datetime(2021, 9, 1)
    cell_date = find_in_range(the_day, range_start_dates)
    self.assertEqual(the_day, cell_date.value)
    self.assertEqual('O', cell_date.column_letter)
    self.assertEqual(160, cell_date.row)

  def test_comment(self):
    wh_xlsx = '9月billing人力-01.xlsx'
    wh_sheet_name = 'Sheet1'
    wh_wb = load_workbook(wh_xlsx)
    wh_ws = wh_wb[wh_sheet_name]
    cell_comment = wh_ws['O35']
    self.assertEqual(8, cell_comment.value)
    self.assertEqual(None, cell_comment.comment)
    # self.assertEqual(' ', cell_comment.comment.author)
    # self.assertEqual('作者:\nextended service 4h', cell_comment.comment.content)
  
  def test_bg_color(self):
    wh_xlsx = '9月billing人力-01.xlsx'
    wh_sheet_name = 'Sheet1'
    wh_wb = load_workbook(wh_xlsx)
    wh_ws = wh_wb[wh_sheet_name]
    color_cell = wh_ws['P1']
    self.assertEqual('FFC4BD97', color_cell.fill.fgColor.rgb)

if __name__ == '__main__':
    # verbosity=*：默认是1；设为0，则不输出每一个用例的执行结果；2-输出详细的执行结果
    unittest.main(verbosity=2)